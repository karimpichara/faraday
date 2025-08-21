from typing import Any
import os
import tempfile

from src.models import db
from src.models.tecnico_supervisor import TecnicoSupervisor


class TecnicoSupervisorService:
    """Service for managing TecnicoSupervisor database operations."""

    def create_tecnico_supervisor(
        self,
        nombre_tecnico: str,
        rut_tecnico: str,
        nombre_supervisor: str,
        id_empresa: int,
    ) -> TecnicoSupervisor:
        """
        Create a new tecnico supervisor record.

        Args:
            nombre_tecnico: Name of the technician
            rut_tecnico: RUT of the technician
            nombre_supervisor: Name of the supervisor
            id_empresa: ID of the empresa

        Returns:
            Created TecnicoSupervisor instance

        Raises:
            RuntimeError: If database operation fails
        """
        try:
            tecnico_supervisor = TecnicoSupervisor(
                nombre_tecnico=nombre_tecnico.strip(),
                rut_tecnico=rut_tecnico.strip(),
                nombre_supervisor=nombre_supervisor.strip(),
                id_empresa=id_empresa,
            )

            db.session.add(tecnico_supervisor)
            db.session.commit()

            return tecnico_supervisor

        except Exception as e:
            db.session.rollback()
            raise RuntimeError(f"Error al crear el técnico supervisor: {str(e)}") from e

    def create_tecnicos_supervisores_bulk(
        self, tecnicos_data: list[dict[str, Any]]
    ) -> dict[str, Any]:
        """
        Create multiple tecnico supervisor records in bulk.

        Args:
            tecnicos_data: List of dictionaries with tecnico data

        Returns:
            Dictionary with operation results

        Raises:
            RuntimeError: If database operation fails
        """
        try:
            created_count = 0
            created_ids = []

            for data in tecnicos_data:
                tecnico_supervisor = TecnicoSupervisor(
                    nombre_tecnico=data["nombre_tecnico"].strip(),
                    rut_tecnico=data["rut_tecnico"].strip(),
                    nombre_supervisor=data["nombre_supervisor"].strip(),
                    id_empresa=data["id_empresa"],
                )

                db.session.add(tecnico_supervisor)
                db.session.flush()  # Get the ID without committing
                created_ids.append(tecnico_supervisor.id)
                created_count += 1

            db.session.commit()

            return {
                "created_count": created_count,
                "created_ids": created_ids,
                "total_count": len(tecnicos_data),
            }

        except Exception as e:
            db.session.rollback()
            raise RuntimeError(
                f"Error al crear los técnicos supervisores: {str(e)}"
            ) from e

    def get_tecnicos_by_empresa(self, id_empresa: int) -> list[TecnicoSupervisor]:
        """
        Get all tecnicos for a specific empresa.

        Args:
            id_empresa: ID of the empresa

        Returns:
            List of TecnicoSupervisor instances
        """
        return TecnicoSupervisor.active_records().filter_by(id_empresa=id_empresa).all()

    def get_tecnico_by_id(self, tecnico_id: int) -> TecnicoSupervisor | None:
        """
        Get a tecnico by ID.

        Args:
            tecnico_id: ID of the tecnico

        Returns:
            TecnicoSupervisor instance or None if not found
        """
        return TecnicoSupervisor.active_records().filter_by(id=tecnico_id).first()

    def get_all_tecnicos_supervisores(self) -> list[TecnicoSupervisor]:
        """
        Get all tecnicos supervisores from all empresas.

        Returns:
            List of all TecnicoSupervisor instances
        """
        return TecnicoSupervisor.active_records().all()

    def process_excel_file(self, file_storage) -> list[dict[str, Any]]:
        """
        Process an uploaded Excel file and extract technician data.

        Args:
            file_storage: FileStorage object from Flask request

        Returns:
            List of dictionaries with technician data

        Raises:
            ValueError: If file processing fails
            RuntimeError: If file operations fail
        """
        try:
            # Validate file
            if not file_storage or file_storage.filename == "":
                raise ValueError("No se ha seleccionado ningún archivo")

            filename = file_storage.filename.lower()
            if not filename.endswith((".xlsx", ".xls")):
                raise ValueError("El archivo debe ser un Excel (.xlsx o .xls)")

            # Save file temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                temp_filepath = temp_file.name
                file_storage.save(temp_filepath)

            try:
                # Process Excel file
                import openpyxl

                wb = openpyxl.load_workbook(temp_filepath, read_only=True)

                # Use first sheet if multiple sheets exist
                sheet_name = wb.sheetnames[0]
                ws = wb[sheet_name]

                tecnicos_data = []

                # Assume first row contains headers
                # Expected columns: nombre_tecnico, rut_tecnico, nombre_supervisor, id_empresa
                headers = []
                for col_num in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col_num).value
                    if cell_value:
                        headers.append(str(cell_value).strip().lower())
                    else:
                        headers.append(f"col_{col_num}")

                # Validate required columns exist
                required_columns = [
                    "nombre_tecnico",
                    "rut_tecnico",
                    "nombre_supervisor",
                    "id_empresa",
                ]
                missing_columns = []
                for req_col in required_columns:
                    if req_col not in headers:
                        missing_columns.append(req_col)

                if missing_columns:
                    raise ValueError(
                        f"Columnas requeridas faltantes: {', '.join(missing_columns)}"
                    )

                # Get column indices
                col_indices = {col: headers.index(col) for col in required_columns}

                # Process data rows (skip header)
                for row_num in range(2, ws.max_row + 1):
                    row_data = {}
                    skip_row = True

                    for col_name, col_idx in col_indices.items():
                        cell_value = ws.cell(row=row_num, column=col_idx + 1).value
                        if cell_value is not None:
                            value = str(cell_value).strip()
                            if value:  # Not empty
                                row_data[col_name] = value
                                if col_name in required_columns:
                                    skip_row = False
                        else:
                            row_data[col_name] = ""

                    # Only add row if it has at least one required field
                    if not skip_row and all(
                        row_data.get(col, "") for col in required_columns
                    ):
                        tecnicos_data.append(row_data)

                wb.close()

                if not tecnicos_data:
                    raise ValueError(
                        "No se encontraron datos válidos en el archivo Excel"
                    )

                return tecnicos_data

            finally:
                # Clean up temporary file
                if os.path.exists(temp_filepath):
                    os.unlink(temp_filepath)

        except ImportError as e:
            raise RuntimeError(
                "La librería openpyxl no está disponible. Instale con: pip install openpyxl"
            ) from e
        except Exception as e:
            if isinstance(e, (ValueError, RuntimeError)):
                raise
            raise RuntimeError(f"Error al procesar el archivo Excel: {str(e)}") from e
